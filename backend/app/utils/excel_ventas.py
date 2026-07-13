"""
Generador de Excel para exportación de ventas (Facturas / Recibos / Ambos).
3 hojas: Encabezado Ventas · Detalle Productos · Formas de Pago
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Paleta de colores por hoja ────────────────────────────────────────────────
_HDR_FILLS = {
    0: PatternFill("solid", fgColor="1E40AF"),   # azul — Encabezado Ventas
    1: PatternFill("solid", fgColor="065F46"),   # verde — Detalle Productos
    2: PatternFill("solid", fgColor="92400E"),   # café  — Formas de Pago
}
_HDR_FONT  = Font(color="FFFFFF", bold=True, size=10)
_ALT_FILL  = PatternFill("solid", fgColor="F1F5F9")
_BORDER    = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)
_ALIGN_CTR = Alignment(horizontal="center", vertical="center")
_ALIGN_L   = Alignment(horizontal="left",   vertical="center", wrap_text=False)


# ── Columnas de cada hoja ─────────────────────────────────────────────────────
COLS_ENCABEZADO = [
    "Tipo", "Numero", "Cedula", "Fecha", "Hora",
    "Cod_Empleado", "Empleado",
    "Total_Efectivo", "T_Credito", "Propina",
    "Anulada", "Id_Resolucion", "Id_Cliente",
]

COLS_DETALLE = [
    "Tipo", "Nro_Pedido", "Fecha", "Nro_Factura",
    "Id_Plato", "Item", "Cantidad", "Valor",
    "Novedad", "Porc_Descuento_Plato", "Porc_Descuento_General",
    "Cambios", "Hora_Plato", "Impuesto",
    "Producto_Personalizado", "Depende",
    # pos_dishes join
    "Plato_Nombre", "Codigo_Producto", "Plato_Valor",
    "Plato_Activo", "Cod_Categoria", "Costo_Producto", "Stock_Minimo",
]

COLS_FORMAS_PAGO = [
    "Tipo", "Item", "Id_Forma_Pago", "Id_Tarjeta",
    "Nro_Factura", "Valor", "Fecha", "Valor_Domicilio", "Nro_Pedido",
    # pos_payment_types join
    "FP_Descripcion", "FP_Valor", "FP_Activo",
]


def _write_sheet(ws, idx: int, title: str, columns: list, rows: list):
    ws.title = title
    ws.freeze_panes = "A2"

    fill = _HDR_FILLS[idx]

    # ── Cabecera ─────────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill   = fill
        cell.font   = _HDR_FONT
        cell.border = _BORDER
        cell.alignment = _ALIGN_CTR
    ws.row_dimensions[1].height = 22

    # ── Datos ─────────────────────────────────────────────────────────────────
    for r_idx, row in enumerate(rows, start=2):
        alt = (r_idx % 2 == 0)
        for c_idx, col_name in enumerate(columns, start=1):
            val = row.get(col_name)
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = _BORDER
            cell.alignment = _ALIGN_L
            if alt:
                cell.fill = _ALT_FILL

    # ── Anchos automáticos ────────────────────────────────────────────────────
    for c_idx, col_name in enumerate(columns, start=1):
        max_len = max(
            len(col_name),
            *(len(str(row.get(col_name, "") or "")) for row in rows[:500])
        ) if rows else len(col_name)
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 3, 40)


def build_ventas_excel(
    rows_enc:  list,   # Sheet 1 — Encabezado Ventas
    rows_det:  list,   # Sheet 2 — Detalle Productos
    rows_fp:   list,   # Sheet 3 — Formas de Pago
) -> bytes:
    wb = Workbook()
    ws1 = wb.active
    _write_sheet(ws1, 0, "Encabezado Ventas",  COLS_ENCABEZADO,  rows_enc)
    _write_sheet(wb.create_sheet(), 1, "Detalle Productos", COLS_DETALLE, rows_det)
    _write_sheet(wb.create_sheet(), 2, "Formas de Pago",    COLS_FORMAS_PAGO, rows_fp)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
